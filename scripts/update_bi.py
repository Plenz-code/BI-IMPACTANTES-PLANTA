"""
update_bi.py
Baixa a planilha do SharePoint via Microsoft Graph API,
processa os dados e atualiza o index.html com os blocos
R_MOT, R_KPI e PROD_PLANTA atualizados.

Variáveis de ambiente necessárias (configurar como GitHub Secrets):
  SP_TENANT_ID      – ID do tenant Azure AD (ex: xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx)
  SP_CLIENT_ID      – Client ID do app registrado no Azure AD
  SP_CLIENT_SECRET  – Secret do app Azure AD
  SP_SITE_URL       – URL do site SharePoint (ex: https://empresa.sharepoint.com/sites/NomeSite)
  SP_FILE_PATH      – Caminho do arquivo dentro do site (ex: /Shared Documents/planilha.xlsx)
"""

import os, sys, json, re, warnings, datetime, io, requests
warnings.filterwarnings('ignore')
from openpyxl import load_workbook

# ── Configuração ─────────────────────────────────────────────────────────────
TENANT_ID     = os.environ['SP_TENANT_ID']
CLIENT_ID     = os.environ['SP_CLIENT_ID']
CLIENT_SECRET = os.environ['SP_CLIENT_SECRET']
SP_SITE_URL   = os.environ['SP_SITE_URL']    # ex: https://empresa.sharepoint.com/sites/Harpia
SP_FILE_PATH  = os.environ['SP_FILE_PATH']   # ex: /Shared Documents/Geral/planilha.xlsx
HTML_FILE     = 'index.html'                 # relativo à raiz do repo

# ── Autenticação Microsoft Graph ──────────────────────────────────────────────
def get_token():
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    r = requests.post(url, data={
        'grant_type':    'client_credentials',
        'client_id':     CLIENT_ID,
        'client_secret': CLIENT_SECRET,
        'scope':         'https://graph.microsoft.com/.default'
    }, timeout=30)
    r.raise_for_status()
    return r.json()['access_token']

def download_file(token):
    from urllib.parse import urlparse
    parsed   = urlparse(SP_SITE_URL)
    hostname = parsed.hostname
    site_path = parsed.path.rstrip('/')

    headers = {'Authorization': f'Bearer {token}'}

    # Resolve o site ID
    site_resp = requests.get(
        f"https://graph.microsoft.com/v1.0/sites/{hostname}:{site_path}",
        headers=headers, timeout=30
    )
    site_resp.raise_for_status()
    site_id = site_resp.json()['id']

    # Baixa o conteúdo do arquivo
    file_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:{SP_FILE_PATH}:/content"
    file_resp = requests.get(file_url, headers=headers, timeout=60)
    file_resp.raise_for_status()
    print(f"✓ Planilha baixada ({len(file_resp.content)//1024} KB)")
    return file_resp.content

# ── Parser da planilha ────────────────────────────────────────────────────────
ABA_MAP = {'Britador': 'brit', 'Linha 1': 'l1', 'Linha 2': 'l2', 'Linha 3': 'l3'}

def fmt_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        d = v.date() if isinstance(v, datetime.datetime) else v
        if 2000 <= d.year <= 2100:
            return d.strftime('%Y-%m-%d')
    return None

def to_mins(v):
    if v is None: return 0
    if isinstance(v, datetime.timedelta): return round(v.total_seconds() / 60)
    if isinstance(v, datetime.time):      return v.hour * 60 + v.minute
    if isinstance(v, datetime.datetime):
        if v.year < 2000: return v.hour * 60 + v.minute
        return round((v.timestamp() % 86400) / 60)
    if isinstance(v, float):
        return round((v - int(v)) * 24 * 60)
    return 0

def clean(v):
    return str(v).replace('\xa0', ' ').strip() if v else ''

def process_workbook(wb_bytes):
    wb = load_workbook(io.BytesIO(wb_bytes), data_only=True)

    _eventos = {k: [] for k in ABA_MAP.values()}
    _has     = {k: {} for k in ABA_MAP.values()}
    _parada  = {k: {} for k in ABA_MAP.values()}

    # ── Paradas por aba ──────────────────────────────────────────────────────
    for sheet, key in ABA_MAP.items():
        if sheet not in wb.sheetnames:
            print(f"  ⚠ Aba '{sheet}' não encontrada, pulando.")
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if isinstance(row[0], str): continue
            data = fmt_date(row[0])
            if not data: continue
            tipo  = clean(row[6])
            equip = clean(row[7])
            desc  = clean(row[8])
            hi    = row[9]
            tp    = row[11]
            conch = row[4] if isinstance(row[4], (int, float)) else 0
            prod  = row[5] if isinstance(row[5], (int, float)) else 0
            t_mins = to_mins(tp)
            if conch > 0 or prod > 0 or tipo or t_mins > 0:
                _has[key][data] = True
            _parada[key].setdefault(data, 0)
            _parada[key][data] += t_mins
            if t_mins > 0 and hi is not None and (tipo or equip or desc):
                _eventos[key].append((data, tipo, equip, desc, t_mins))

    # ── Produção da planta (balança das linhas + Resumo) ────────────────────
    balanca     = {k: {} for k in ABA_MAP.values()}
    prod_linha  = {k: {} for k in ABA_MAP.values()}
    material_tot = {k: {} for k in ABA_MAP.values()}

    for sheet, key in ABA_MAP.items():
        if sheet not in wb.sheetnames: continue
        ws = wb[sheet]
        ultima_data = None
        turno_soma  = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            c0   = row[0]
            data = fmt_date(c0)
            if data:
                ultima_data = data
                conch = row[4] if isinstance(row[4], (int, float)) else 0
                prod  = row[5] if isinstance(row[5], (int, float)) else 0
                mat   = clean(row[3])
                prod_linha[key].setdefault(data, {'conch': 0, 'prod': 0})
                prod_linha[key][data]['conch'] += conch
                prod_linha[key][data]['prod']  += prod
                if mat and conch > 0:
                    material_tot[key][mat] = material_tot[key].get(mat, 0) + conch
            elif isinstance(c0, str):
                cu  = c0.upper()
                bal = row[13] if (len(row) > 13 and isinstance(row[13], (int, float))) else 0
                if 'TURNO' in cu:
                    turno_soma += bal
                elif 'TOTAL GLOBAL' in cu:
                    if ultima_data:
                        valor = bal if bal > 0 else turno_soma
                        if valor > 0:
                            balanca[key][ultima_data] = valor
                    turno_soma = 0

    # Complementa com aba Resumo se existir
    if 'Resumo' in wb.sheetnames:
        ws_r = wb['Resumo']
        for row in ws_r.iter_rows(min_row=4, max_row=33, values_only=True):
            dr = row[1]
            if not isinstance(dr, (datetime.datetime, datetime.date)): continue
            d_obj = dr.date() if isinstance(dr, datetime.datetime) else dr
            data  = d_obj.strftime('%Y-%m-%d')
            for i, k in enumerate(['brit', 'l1', 'l2', 'l3'], start=2):
                v = row[i]
                if isinstance(v, (int, float)) and v > 0 and data not in balanca[k]:
                    balanca[k][data] = float(v)

    # ── Builds ───────────────────────────────────────────────────────────────
    def build_daily(pm, hm):
        out = {}
        for d, p in sorted(pm.items()):
            if not hm.get(d): continue
            hs = round(min(24, p / 60), 1)
            out[d] = {'prod': round(max(0, 24 - hs), 1), 'parada': hs}
        return out

    def build_mots(evs):
        bd = {}
        for data, tipo, equip, desc, mins in evs:
            bd.setdefault(data, [])
            bd[data].append([f"{tipo} \u2225 {equip} \u2225 {desc}", round(mins / 60, 2)])
        for d in bd:
            bd[d].sort(key=lambda x: -x[1])
        return bd

    def build_rkpi(daily):
        out = {}
        for d, v in daily.items():
            tb = v['prod'] + v['parada']
            out[d] = {
                'hp': v['prod'], 'hs': v['parada'],
                'df': round(v['prod'] / tb * 100, 1) if tb > 0 else 0
            }
        return out

    results = {}
    for key in ABA_MAP.values():
        daily = build_daily(_parada[key], _has[key])
        results[key] = {
            'daily': daily,
            'mots':  build_mots(_eventos[key]),
            'rkpi':  build_rkpi(daily)
        }
        n = sum(len(v) for v in results[key]['mots'].values())
        ultimo = max(daily.keys()) if daily else '—'
        print(f"  {key}: {len(daily)} dias, {n} ocorrências, último: {ultimo}")

    # Monta resumo de produção
    todas_datas = sorted(set(d for k in balanca for d in balanca[k]))
    resumo = []
    for data in todas_datas:
        b = balanca['brit'].get(data, 0)
        v1 = balanca['l1'].get(data, 0)
        v2 = balanca['l2'].get(data, 0)
        v3 = balanca['l3'].get(data, 0)
        resumo.append({
            'data': data, 'dia': int(data[8:10]),
            'brit': b, 'l1': v1, 'l2': v2, 'l3': v3,
            'total': v1 + v2 + v3
        })

    prod_planta = {'resumo': resumo, 'prod_linha': prod_linha, 'material_tot': material_tot}
    wb.close()
    return results, prod_planta

# ── Atualização do HTML ───────────────────────────────────────────────────────
def update_html(results, prod_planta):
    with open(HTML_FILE, encoding='utf-8') as f:
        html = f.read()

    def js(o): return json.dumps(o, ensure_ascii=False, separators=(',', ':'))

    rmot = {k: v['mots'] for k, v in results.items()}
    rkpi = {k: v['rkpi'] for k, v in results.items()}

    new_rmot_block = (
        f"let R_MOT = {{\n"
        f"  brit: {js(rmot['brit'])},\n"
        f"  l1: {js(rmot['l1'])},\n"
        f"  l2: {js(rmot['l2'])},\n"
        f"  l3: {js(rmot['l3'])}\n"
        f"}};"
    )
    new_rkpi_block = (
        f"let R_KPI = {{\n"
        f"  brit: {js(rkpi['brit'])},\n"
        f"  l1: {js(rkpi['l1'])},\n"
        f"  l2: {js(rkpi['l2'])},\n"
        f"  l3: {js(rkpi['l3'])}\n"
        f"}};"
    )
    new_pp_block = f"let PROD_PLANTA = {js(prod_planta)};"

    # Substitui blocos no HTML (regex para tolerar variação de espaços)
    html = re.sub(r'(let|const)\s+R_MOT\s*=\s*\{[\s\S]*?\};', new_rmot_block, html, count=1)
    html = re.sub(r'(let|const)\s+R_KPI\s*=\s*\{[\s\S]*?\};', new_rkpi_block, html, count=1)
    html = re.sub(r'(let|const)\s+PROD_PLANTA\s*=\s*\{[\s\S]*?\};', new_pp_block, html, count=1)

    with open(HTML_FILE, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"✓ {HTML_FILE} atualizado")

# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print("🔑 Autenticando com Microsoft Graph...")
    token = get_token()

    print("📥 Baixando planilha do SharePoint...")
    xlsx_bytes = download_file(token)

    print("⚙  Processando dados...")
    results, prod_planta = process_workbook(xlsx_bytes)

    print("📝 Atualizando index.html...")
    update_html(results, prod_planta)

    print("✅ Concluído!")

