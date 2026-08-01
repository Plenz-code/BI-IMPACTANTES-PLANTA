"""
update_bi_tudo.py
1. Baixa planilha do Dropbox (link publico)
2. Processa os dados
3. Atualiza o index.html com auto-inicializacao
"""

import os, sys, json, re, warnings, datetime, zipfile, urllib.request
warnings.filterwarnings('ignore')
from openpyxl import load_workbook

XLSX_FILE   = 'planilha.xlsx'
HTML_FILE   = 'index.html'
DROPBOX_URL = os.environ.get('DROPBOX_URL', '')
ABA_MAP     = {'Britador':'brit','Linha 1':'l1','Linha 2':'l2','Linha 3':'l3'}

# ── 1. BAIXAR DO DROPBOX ──────────────────────────────────────────────────────
def normaliza_dropbox_url(url):
    """Garante que o link do Dropbox force download direto do arquivo bruto."""
    if not url:
        return url
    # Links antigos (?dl=0) -> força dl=1
    if 'dropbox.com' in url and 'dl=' in url:
        url = re.sub(r'dl=0(&|$)', r'dl=1\1', url)
        if 'dl=1' not in url:
            url += ('&' if '?' in url else '?') + 'dl=1'
    # Links sem parametro dl e sem raw -> adiciona dl=1
    elif 'dropbox.com' in url and 'dl=' not in url and 'dropboxusercontent' not in url:
        url += ('&' if '?' in url else '?') + 'dl=1'
    return url

if DROPBOX_URL:
    DROPBOX_URL = normaliza_dropbox_url(DROPBOX_URL)
    print(f"Baixando do Dropbox... ({DROPBOX_URL[:60]}...)")
    try:
        req = urllib.request.Request(DROPBOX_URL, headers={'User-Agent':'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = resp.read()
            content_type = resp.headers.get('Content-Type', '?')
        print(f"  Content-Type: {content_type} | Tamanho: {len(data)} bytes")

        # xlsx (zip) sempre comeca com os bytes PK (0x50 0x4B)
        if not data.startswith(b'PK'):
            preview = data[:300].decode('utf-8', errors='replace')
            print("Arquivo baixado NAO e um xlsx valido (nao comeca com assinatura ZIP/PK).")
            print(f"Content-Type recebido: {content_type}")
            print(f"Primeiros bytes (preview): {preview}")
            raise Exception(
                "O Dropbox retornou uma pagina/HTML em vez do arquivo. "
                "O link provavelmente mudou (ex: pasta do SharePoint foi alterada, "
                "gerando um novo link de compartilhamento). "
                "Atualize o secret DROPBOX_URL no GitHub com o link atual e certifique-se "
                "de que termina com '?dl=1' (download direto)."
            )
        if len(data) < 10000:
            raise Exception(f"Arquivo muito pequeno ({len(data)} bytes) — pode estar corrompido ou vazio")

        with open(XLSX_FILE, 'wb') as f:
            f.write(data)
        print(f"✓ Baixado do Dropbox ({len(data)//1024} KB)")
    except Exception as e:
        print(f"Dropbox falhou: {e}")
        if not os.path.exists(XLSX_FILE):
            print("Nenhum arquivo disponivel — abortando")
            sys.exit(1)
        print("Usando planilha.xlsx existente no repositorio")
else:
    print("DROPBOX_URL nao configurado — usando planilha.xlsx do repositorio")
    if not os.path.exists(XLSX_FILE):
        print("planilha.xlsx nao encontrada — abortando")
        sys.exit(1)

# ── Valida o arquivo ──────────────────────────────────────────────────────────
try:
    with zipfile.ZipFile(XLSX_FILE, 'r') as z:
        z.testzip()
    print(f"✓ planilha.xlsx valida ({os.path.getsize(XLSX_FILE)//1024} KB)")
except Exception as e:
    print(f"Arquivo invalido: {e}")
    sys.exit(1)

# ── 2. PROCESSAR ─────────────────────────────────────────────────────────────
def fmt_date(v):
    if isinstance(v,(datetime.datetime,datetime.date)):
        d=v.date() if isinstance(v,datetime.datetime) else v
        if 2000<=d.year<=2100: return d.strftime('%Y-%m-%d')
    return None

def to_mins(v):
    if v is None: return 0
    if isinstance(v,datetime.timedelta): return round(v.total_seconds()/60)
    if isinstance(v,datetime.time): return v.hour*60+v.minute
    if isinstance(v,datetime.datetime):
        if v.year<2000: return v.hour*60+v.minute
        return round((v.timestamp()%86400)/60)
    if isinstance(v,float): return round((v-int(v))*24*60)
    return 0

def clean(v): return str(v).replace('\xa0',' ').strip() if v else ''

wb = load_workbook(XLSX_FILE, data_only=True)
print(f"Planilha carregada: {XLSX_FILE}")

_eventos={k:[] for k in ABA_MAP.values()}
_has={k:{} for k in ABA_MAP.values()}
_parada={k:{} for k in ABA_MAP.values()}
balanca={k:{} for k in ABA_MAP.values()}
prod_linha={k:{} for k in ABA_MAP.values()}
material_tot={k:{} for k in ABA_MAP.values()}

for sheet,key in ABA_MAP.items():
    if sheet not in wb.sheetnames: continue
    ws=wb[sheet]
    ultima_data=None; turno_soma=0
    for row in ws.iter_rows(min_row=2,values_only=True):
        c0=row[0]; data=fmt_date(c0)
        if data:
            ultima_data=data
            tipo=clean(row[6]); equip=clean(row[7]); desc=clean(row[8])
            hi=row[9]; tp=row[11]
            conch=row[4] if isinstance(row[4],(int,float)) else 0
            prod=row[5] if isinstance(row[5],(int,float)) else 0
            mat=clean(row[3]); t_mins=to_mins(tp)
            if conch>0 or prod>0 or tipo or t_mins>0: _has[key][data]=True
            _parada[key].setdefault(data,0); _parada[key][data]+=t_mins
            if t_mins>0 and hi is not None and (tipo or equip or desc):
                _eventos[key].append((data,tipo,equip,desc,t_mins))
            prod_linha[key].setdefault(data,{'conch':0,'prod':0})
            prod_linha[key][data]['conch']+=conch; prod_linha[key][data]['prod']+=prod
            if mat and conch>0: material_tot[key][mat]=material_tot[key].get(mat,0)+conch
        elif isinstance(c0,str):
            cu=c0.upper(); bal=row[13] if(len(row)>13 and isinstance(row[13],(int,float))) else 0
            if 'TURNO' in cu: turno_soma+=bal
            elif 'TOTAL GLOBAL' in cu:
                if ultima_data:
                    valor=bal if bal>0 else turno_soma
                    if valor>0: balanca[key][ultima_data]=valor
                turno_soma=0

if 'Resumo' in wb.sheetnames:
    for row in wb['Resumo'].iter_rows(min_row=4,max_row=33,values_only=True):
        dr=row[1]
        if not isinstance(dr,(datetime.datetime,datetime.date)): continue
        d=dr.date() if isinstance(dr,datetime.datetime) else dr
        dt=d.strftime('%Y-%m-%d')
        for i,k in enumerate(['brit','l1','l2','l3'],start=2):
            v=row[i]
            if isinstance(v,(int,float)) and v>0 and dt not in balanca[k]: balanca[k][dt]=float(v)
wb.close()

def build_daily(pm,hm):
    out={}
    for d,p in sorted(pm.items()):
        if not hm.get(d): continue
        hs=round(min(24,p/60),1); out[d]={'prod':round(max(0,24-hs),1),'parada':hs}
    return out

def build_mots(evs):
    bd={}
    for data,tipo,equip,desc,mins in evs:
        bd.setdefault(data,[]); bd[data].append([f"{tipo} \u2225 {equip} \u2225 {desc}",round(mins/60,2)])
    for d in bd: bd[d].sort(key=lambda x:-x[1])
    return bd

def build_rkpi(daily):
    return {d:{'hp':v['prod'],'hs':v['parada'],'df':round(v['prod']/(v['prod']+v['parada'])*100,1) if(v['prod']+v['parada'])>0 else 0} for d,v in daily.items()}

results={}
for key in ABA_MAP.values():
    daily=build_daily(_parada[key],_has[key])
    mots=build_mots(_eventos[key])
    results[key]={'mots':mots,'rkpi':build_rkpi(daily)}
    print(f"  {key}: {len(daily)} dias, {sum(len(v) for v in mots.values())} ocorrencias")

todas_datas=sorted(set(d for k in balanca for d in balanca[k]))
resumo=[{'data':dt,'dia':int(dt[8:10]),'brit':balanca['brit'].get(dt,0),'l1':balanca['l1'].get(dt,0),'l2':balanca['l2'].get(dt,0),'l3':balanca['l3'].get(dt,0),'total':balanca['l1'].get(dt,0)+balanca['l2'].get(dt,0)+balanca['l3'].get(dt,0)} for dt in todas_datas]
prod_planta={'resumo':resumo,'prod_linha':prod_linha,'material_tot':material_tot}

# ── 3. ATUALIZAR HTML ─────────────────────────────────────────────────────────
print(f"Atualizando {HTML_FILE}...")
with open(HTML_FILE,encoding='utf-8') as f: html=f.read()

rmot={k:v['mots'] for k,v in results.items()}
rkpi={k:v['rkpi'] for k,v in results.items()}
js=lambda o:json.dumps(o,ensure_ascii=False,separators=(',',':'))

html=re.sub(r'(let|const)\s+R_MOT\s*=\s*\{[\s\S]*?\};',f"let R_MOT = {{\n  brit: {js(rmot['brit'])},\n  l1: {js(rmot['l1'])},\n  l2: {js(rmot['l2'])},\n  l3: {js(rmot['l3'])}\n}};",html,count=1)
html=re.sub(r'(let|const)\s+R_KPI\s*=\s*\{[\s\S]*?\};',f"let R_KPI = {{\n  brit: {js(rkpi['brit'])},\n  l1: {js(rkpi['l1'])},\n  l2: {js(rkpi['l2'])},\n  l3: {js(rkpi['l3'])}\n}};",html,count=1)
html=re.sub(r'(let|const)\s+PROD_PLANTA\s*=\s*\{[\s\S]*?\};',f"let PROD_PLANTA = {js(prod_planta)};",html,count=1)

AUTO_INIT="""<!-- AUTO_INIT_START -->
<script>
document.addEventListener('DOMContentLoaded',function(){
  if(!R_KPI||!R_KPI.l1||!Object.keys(R_KPI.l1).length)return;
  ['brit','l1','l2','l3'].forEach(function(k){
    DADOS[k+'_daily']={};
    Object.entries(R_KPI[k]||{}).forEach(function(e){DADOS[k+'_daily'][e[0]]={prod:e[1].hp,parada:e[1].hs};});
    var acc={};
    Object.values(R_MOT[k]||{}).forEach(function(l){(l||[]).forEach(function(i){acc[i[0]]=(acc[i[0]]||0)+i[1];});});
    DADOS[k].top_motivos=Object.entries(acc).map(function(e){return[e[0],e[1]];}).sort(function(a,b){return b[1]-a[1];}).slice(0,14);
  });
  setTimeout(function(){var b=document.getElementById('btnDash');if(b&&typeof goTab==='function')goTab('dashboard',b);},150);
});
</script>
<!-- AUTO_INIT_END -->"""

html=re.sub(r'<!-- AUTO_INIT_START -->[\s\S]*?<!-- AUTO_INIT_END -->','',html)
html=html.replace('</body>',AUTO_INIT+'\n</body>',1)

with open(HTML_FILE,'w',encoding='utf-8') as f: f.write(html)
print("✓ index.html atualizado!")
